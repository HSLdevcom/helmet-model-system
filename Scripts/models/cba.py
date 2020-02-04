from openpyxl import Workbook, load_workbook
import sys
import os
import openmatrix as omx
import numpy
import pandas
import re
import parameters as param


def do_cba(alt_0, alt_1, year, excelfile):
    """Runs CBA and writes the results to excel file.

    Parameters
    ----------
    alt_0 : str
        Name of do-nothing scenario, for which 
        forecast results are available in Results folder
    alt_1 : str
        Name of project scenario, for which 
        forecast results are available in Results folder
    year : int
        The evaluation year (1 or 2)
    excelfile : str
        Path to excel file where results will be written
    """
    mile_diff = miles(alt_1) - miles(alt_0)
    transit_mile_diff = transit_miles(alt_1) - transit_miles(alt_0)
    revenues = {
        "car": {},
        "transit": {},
    }
    gains = dict.fromkeys(param.transport_classes)
    for transport_class in gains:
        gains[transport_class] = {}
    for tp in param.emme_scenario:
        ve1 = read_scen(alt_1, tp)
        ve0 = read_scen(alt_0, tp)
        revenues["transit"][tp] = 0
        for transit_class in ("transit_work", "transit_leisure"):
            revenues["transit"][tp] += calc_revenue(
                ve0[transit_class], ve1[transit_class])
        revenues["car"][tp] = 0
        for ass_mode in param.assignment_mode:
            revenues["car"][tp] += calc_revenue(ve0[ass_mode], ve1[ass_mode])
        print "Revenues " + tp + " calculated"
        for transport_class in gains:
            gains[transport_class][tp] = calc_gains(
                ve0[transport_class], ve1[transport_class])
        print "Gains " + tp + " calculated"
    wb = load_workbook(excelfile)
    if year == 1:
        write_results_1(wb, mile_diff, transit_mile_diff, revenues, gains)
    elif year == 2:
        write_results_2(wb, mile_diff, transit_mile_diff, revenues, gains)
    else:
        print "Evaluation year must be either 1 or 2"
    wb.save("..\\Results\\cba_" + alt_1 + ".xlsx")

def read_scen(scenario, time_period):
    """Read travel cost and demand data for scenario from files"""
    script_dir = os.path.dirname(os.path.realpath(__file__))
    project_dir = os.path.join(script_dir, "..", "..")
    path = os.path.join(project_dir, "Matrices", scenario)
    files = dict.fromkeys(["demand", "time", "cost", "dist"])
    for mtx_type in files:
        file_name = mtx_type + '_' + time_period + ".omx"
        file_path = os.path.join(path, file_name)
        files[mtx_type] = omx.open_file(file_path)
    matrices = {}
    for transport_class in param.transport_classes:
        matrices[transport_class] = {}
        for mtx_type in files:
            if mtx_type != "demand":
                mtx_label = transport_class.split('_')[0]
                if mtx_label == "transit" or mtx_label == "bike":
                    ass_class = mtx_label
                else:
                    ass_class = transport_class
            else:
                ass_class = transport_class
            if mtx_label == "bike" and mtx_type == "cost":
                matrices[transport_class][mtx_type] = 0
            else:
                matrices[transport_class][mtx_type] = numpy.array(files[mtx_type][ass_class])
    for mtx_type in files:
        files[mtx_type].close()
    print "Files read"
    return matrices    

def calc_revenue(ve0, ve1):
    """Calculate difference in producer revenue between scenarios ve1 and ve0"""
    revenue = 0
    demand_change = ve1["demand"] - ve0["demand"]
    cost_change = ve1["cost"] - ve0["cost"]
    revenue += (ve1["cost"] * demand_change)[demand_change >= 0].sum()
    revenue += (cost_change * ve0["demand"])[demand_change >= 0].sum()
    revenue += (ve0["cost"] * demand_change)[demand_change < 0].sum()
    revenue += (cost_change * ve1["demand"])[demand_change < 0].sum()
    return revenue

def calc_cost_gains(ve0, ve1):
    """Calculate difference in consumer surplus between scenarios ve1 and ve0"""
    gains = {"existing": 0, "additional": 0}
    gain = ve1["cost"] - ve0["cost"]
    demand_change = ve1["demand"] - ve0["demand"]
    gains["existing"] += (ve0["demand"] * gain)[demand_change >= 0].sum()
    gains["additional"] += 0.5 * (demand_change * gain)[demand_change >= 0].sum()
    gains["existing"] += (ve1["demand"] * gain)[demand_change < 0].sum()
    gains["additional"] -= 0.5 * (demand_change * gain)[demand_change < 0].sum()
    return gains

def calc_gains(ve0, ve1):
    """Calculate time, distance and cost gains"""
    gains = {}
    gains["time"] = calc_cost_gains(
        {
            "cost": ve0["time"],
            "demand": ve0["demand"],
        },
        {
            "cost": ve1["time"],
            "demand": ve1["demand"],
        },
    )
    gains["dist"] = calc_cost_gains(
        {
            "cost": ve0["dist"],
            "demand": ve0["demand"],
        },
        {
            "cost": ve1["dist"],
            "demand": ve1["demand"],
        },
    )
    gains["cost"] = calc_cost_gains(ve0, ve1)
    return gains

def miles(scenario):
    """Read scenario data from files"""
    script_dir = os.path.dirname(os.path.realpath(__file__))
    project_dir = os.path.join(script_dir, "..", "..")
    data_dir = os.path.join(project_dir, "Results", scenario)
    data_dir = os.path.abspath(data_dir)
    filename = os.path.join(data_dir, "vehicle_kms.txt")
    data = pandas.read_csv(filename, delim_whitespace=True)
    return data

def transit_miles(scenario):
    """Read scenario data from files"""
    script_dir = os.path.dirname(os.path.realpath(__file__))
    project_dir = os.path.join(script_dir, "..", "..")
    data_dir = os.path.join(project_dir, "Results", scenario)
    data_dir = os.path.abspath(data_dir)
    filename = os.path.join(data_dir, "transit_kms.txt")
    data = pandas.read_csv(filename, delim_whitespace=True)
    return data

def write_results_1(wb, miles, transit_miles, revenues, gains):
    """Write results for year 1"""
    ws = wb.get_sheet_by_name("ha_tyo")
    write_gains_1(ws, gains["car_work"])
    ws = wb.get_sheet_by_name("ha_muu")
    write_gains_1(ws, gains["car_leisure"])
    ws = wb.get_sheet_by_name("jl_tyo")
    write_gains_1(ws, gains["transit_work"])
    ws = wb.get_sheet_by_name("jl_muu")
    write_gains_1(ws, gains["transit_leisure"])
    ws = wb.get_sheet_by_name("pp_tyo")
    write_gains_1(ws, gains["bike_work"])
    ws = wb.get_sheet_by_name("pp_muu")
    write_gains_1(ws, gains["bike_leisure"])
    ws = wb.get_sheet_by_name("ka")
    write_gains_1(ws, gains["truck"])
    ws = wb.get_sheet_by_name("yhd")
    write_gains_1(ws, gains["trailer_truck"])
    ws = wb.get_sheet_by_name("pa")
    write_gains_1(ws, gains["van"])
    ws = wb.get_sheet_by_name("Ulkoisvaikutukset")
    ws["I19"] = miles["car"][1]
    ws["J19"] = miles["car"][2]
    ws["K19"] = miles["car"][3]
    ws["L19"] = miles["car"][4]
    ws["M19"] = miles["car"][5]
    ws["I20"] = miles["van"][1]
    ws["J20"] = miles["van"][2]
    ws["K20"] = miles["van"][3]
    ws["L20"] = miles["van"][4]
    ws["M20"] = miles["van"][5]
    ws["I21"] = miles["truck"][1]
    ws["J21"] = miles["truck"][2]
    ws["K21"] = miles["truck"][3]
    ws["L21"] = miles["truck"][4]
    ws["M21"] = miles["truck"][5]
    ws["I22"] = miles["trailer_truck"][1]
    ws["J22"] = miles["trailer_truck"][2]
    ws["K22"] = miles["trailer_truck"][3]
    ws["L22"] = miles["trailer_truck"][4]
    ws["M22"] = miles["trailer_truck"][5]
    ws = wb.get_sheet_by_name("Tuottajahyodyt")
    ws["S8"] = transit_miles["dist"]["bus"]
    ws["S9"] = transit_miles["dist"]["trunk"]
    ws["S10"] = transit_miles["dist"]["tram"]
    ws["S11"] = transit_miles["dist"]["metro"]
    ws["S12"] = transit_miles["dist"]["train"]
    ws["T8"] = transit_miles["time"]["bus"]
    ws["T9"] = transit_miles["time"]["trunk"]
    ws["T10"] = transit_miles["time"]["tram"]
    ws["T11"] = transit_miles["time"]["metro"]
    ws["T12"] = transit_miles["time"]["train"]
    ws["B43"] = revenues["transit"]["aht"]
    ws["C43"] = revenues["transit"]["pt"]
    ws["D43"] = revenues["transit"]["iht"]
    ws = wb.get_sheet_by_name("Julkistaloudelliset")
    ws["F8"] = revenues["car"]["aht"]
    ws["G8"] = revenues["car"]["pt"]
    ws["H8"] = revenues["car"]["iht"]

def write_gains_1(ws, gains):
    ws["B9"] = gains["aht"]["time"]["existing"]
    ws["B10"] = gains["aht"]["time"]["additional"]
    ws["C9"] = gains["pt"]["time"]["existing"]
    ws["C10"] = gains["pt"]["time"]["additional"]
    ws["D9"] = gains["iht"]["time"]["existing"]
    ws["D10"] = gains["iht"]["time"]["additional"]
    ws["B22"] = gains["aht"]["dist"]["existing"]
    ws["B23"] = gains["aht"]["dist"]["additional"]
    ws["C22"] = gains["pt"]["dist"]["existing"]
    ws["C23"] = gains["pt"]["dist"]["additional"]
    ws["D22"] = gains["iht"]["dist"]["existing"]
    ws["D23"] = gains["iht"]["dist"]["additional"]
    ws["B37"] = gains["aht"]["cost"]["existing"]
    ws["B38"] = gains["aht"]["cost"]["additional"]
    ws["C37"] = gains["pt"]["cost"]["existing"]
    ws["C38"] = gains["pt"]["cost"]["additional"]
    ws["D37"] = gains["iht"]["cost"]["existing"]
    ws["D38"] = gains["iht"]["cost"]["additional"]

def write_results_2(wb, miles, transit_miles, revenues, gains):
    """Write results for year 2"""
    ws = wb.get_sheet_by_name("ha_tyo")
    write_gains_2(ws, gains["car_work"])
    ws = wb.get_sheet_by_name("ha_muu")
    write_gains_2(ws, gains["car_leisure"])
    ws = wb.get_sheet_by_name("jl_tyo")
    write_gains_2(ws, gains["transit_work"])
    ws = wb.get_sheet_by_name("jl_muu")
    write_gains_2(ws, gains["transit_leisure"])
    ws = wb.get_sheet_by_name("pp_tyo")
    write_gains_2(ws, gains["bike_work"])
    ws = wb.get_sheet_by_name("pp_muu")
    write_gains_2(ws, gains["bike_leisure"])
    ws = wb.get_sheet_by_name("ka")
    write_gains_2(ws, gains["truck"])
    ws = wb.get_sheet_by_name("yhd")
    write_gains_2(ws, gains["trailer_truck"])
    ws = wb.get_sheet_by_name("pa")
    write_gains_2(ws, gains["van"])
    ws = wb.get_sheet_by_name("Ulkoisvaikutukset")
    ws["I32"] = miles["car"][1]
    ws["J32"] = miles["car"][2]
    ws["K32"] = miles["car"][3]
    ws["L32"] = miles["car"][4]
    ws["M32"] = miles["car"][5]
    ws["I33"] = miles["van"][1]
    ws["J33"] = miles["van"][2]
    ws["K33"] = miles["van"][3]
    ws["L33"] = miles["van"][4]
    ws["M33"] = miles["van"][5]
    ws["I34"] = miles["truck"][1]
    ws["J34"] = miles["truck"][2]
    ws["K34"] = miles["truck"][3]
    ws["L34"] = miles["truck"][4]
    ws["M34"] = miles["truck"][5]
    ws["I35"] = miles["trailer_truck"][1]
    ws["J35"] = miles["trailer_truck"][2]
    ws["K35"] = miles["trailer_truck"][3]
    ws["L35"] = miles["trailer_truck"][4]
    ws["M35"] = miles["trailer_truck"][5]
    ws = wb.get_sheet_by_name("Kayttajahyodyt")
    ws = wb.get_sheet_by_name("Tuottajahyodyt")
    ws["S16"] = transit_miles["dist"]["bus"]
    ws["S17"] = transit_miles["dist"]["trunk"]
    ws["S18"] = transit_miles["dist"]["tram"]
    ws["S19"] = transit_miles["dist"]["metro"]
    ws["S20"] = transit_miles["dist"]["train"]
    ws["T16"] = transit_miles["time"]["bus"]
    ws["T17"] = transit_miles["time"]["trunk"]
    ws["T18"] = transit_miles["time"]["tram"]
    ws["T19"] = transit_miles["time"]["metro"]
    ws["T20"] = transit_miles["time"]["train"]
    ws["B46"] = revenues["transit"]["aht"]
    ws["C46"] = revenues["transit"]["pt"]
    ws["D46"] = revenues["transit"]["iht"]
    ws = wb.get_sheet_by_name("Julkistaloudelliset")
    ws["F13"] = revenues["car"]["aht"]
    ws["G13"] = revenues["car"]["pt"]
    ws["H13"] = revenues["car"]["iht"]

def write_gains_2(ws, gains):
    ws["B14"] = gains["aht"]["time"]["existing"]
    ws["B15"] = gains["aht"]["time"]["additional"]
    ws["C14"] = gains["pt"]["time"]["existing"]
    ws["C15"] = gains["pt"]["time"]["additional"]
    ws["D14"] = gains["iht"]["time"]["existing"]
    ws["D15"] = gains["iht"]["time"]["additional"]
    ws["B27"] = gains["aht"]["dist"]["existing"]
    ws["B28"] = gains["aht"]["dist"]["additional"]
    ws["C27"] = gains["pt"]["dist"]["existing"]
    ws["C28"] = gains["pt"]["dist"]["additional"]
    ws["D27"] = gains["iht"]["dist"]["existing"]
    ws["D28"] = gains["iht"]["dist"]["additional"]
    ws["B42"] = gains["aht"]["cost"]["existing"]
    ws["B43"] = gains["aht"]["cost"]["additional"]
    ws["C42"] = gains["pt"]["cost"]["existing"]
    ws["C43"] = gains["pt"]["cost"]["additional"]
    ws["D42"] = gains["iht"]["cost"]["existing"]
    ws["D43"] = gains["iht"]["cost"]["additional"]


do_cba(sys.argv[1], sys.argv[2], int(sys.argv[3]), sys.argv[4])
# do_cba("2030_test", "2030_test", 1, "..\\CBA_kehikko.xlsx")