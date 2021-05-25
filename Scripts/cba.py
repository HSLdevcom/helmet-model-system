from openpyxl import load_workbook
import os
import numpy
import pandas
from argparse import ArgumentError, ArgumentParser

from utils.config import Config
import utils.log as log
import parameters.assignment as param
import parameters.zone as zone_param
from datahandling.matrixdata import MatrixData


SCRIPT_DIR = os.path.dirname(os.path.realpath(__file__))

VEHICLE_KMS_FILE = "vehicle_kms_vdfs.txt"
TRANSIT_KMS_FILE = "transit_kms.txt"

TRANSIT_TRIPS_PER_MONTH = {
    "work_capital_region": 60,
    "work_surrounding": 44,
    "leisure": 30,
}

TRANSIT_AGGREGATIONS = {
    "bus": ("HSL-bussi", "ValluVakio", "ValluPika"),
    "train": ("HSL-juna", "muu_juna"),
    "tram": ("ratikka", "pikaratikk"),
}

TRANSLATIONS = {
    "car_work": "ha_tyo",
    "car_leisure": "ha_muu",
    "transit_work": "jl_tyo",
    "transit_leisure": "jl_muu",
    "bike_work": "pp_tyo",
    "bike_leisure": "pp_muu",
    "truck": "ka",
    "trailer_truck": "yhd",
    "van": "pa",
}

CELL_INDICES = {
    "gains": {
        "cols": {
            "aht": 'B',
            "pt": 'C',
            "iht": 'D',
        },
        "rows": {
            # Different rows for the two years
            # One row for existing users, one row for additional users
            1: {
                "time": ("9", "10"),
                "dist": ("22", "23"),
                "cost": ("37", "38"),
            },
            2: {
                "time": ("14", "15"),
                "dist": ("27", "28"),
                "cost": ("42", "43"),
            },
        },
    },
    "transit_revenue": {
        "rows": {
            1: "43",
            2: "46",
        },
    },
    "car_revenue": {
        "cols": {
            "aht": 'F',
            "pt": 'G',
            "iht": 'H',
        },
        "rows": {
            1: "8",
            2: "13",
        },
    },
    "car_miles": {
        "cols": {
            # Column index for each volume-delay function
            1: 'I',
            2: 'J',
            3: 'K',
            4: 'L',
            5: 'M',
        },
        "rows": {
            1: {
                "car": "19",
                "van": "20",
                "truck": "21",
                "trailer_truck": "22",
            },
            2: {
                "car": "32",
                "van": "33",
                "truck": "34",
                "trailer_truck": "35",
            },
        },
    },
    "transit_miles": {
        "cols": {
            "dist": 'S',
            "time": 'T',
        },
        "rows": {
            1: {
                "bus": "8",
                "HSL-runkob": "9",
                "tram": "10",
                "metro": "11",
                "train": "12",
            },
            2: {
                "bus": "16",
                "HSL-runkob": "17",
                "tram": "18",
                "metro": "19",
                "train": "20",
            },
        },
    },
}

def run_cost_benefit_analysis(scenario_0, scenario_1, year, workbook):
    """Runs CBA and writes the results to excel file.

    Parameters
    ----------
    scenario_0 : str
        Name of do-nothing scenario, for which 
        forecast results are available in Results folder
    scenario_1 : str
        Name of project scenario, for which 
        forecast results are available in Results folder
    year : int
        The evaluation year (1 or 2)
    results_directory : str
        Path to where "scenario_name/Matrices" result folder exists
    workbook : openpyxl.WorkBook
        The excel workbook where to save results
    """
    if year not in (1, 2):
        raise ArgumentError("Evaluation year must be either 1 or 2")
    log.info("Analyse year {}...".format(year))

    # Calculate mile differences
    mile_diff = read_miles(scenario_1) - read_miles(scenario_0)
    mile_diff["car"] = mile_diff["car_work"] + mile_diff["car_leisure"]
    ws = workbook["Ulkoisvaikutukset"]
    cols = CELL_INDICES["car_miles"]["cols"]
    rows = CELL_INDICES["car_miles"]["rows"][year]
    for mode in rows:
        for vdf in cols:
            ws[cols[vdf]+rows[mode]] = mile_diff[mode][vdf]

    # Calculate transit mile differences
    transit_mile_diff = (read_transit_miles(scenario_1)
                         - read_transit_miles(scenario_0))
    for mode in TRANSIT_AGGREGATIONS:
        transit_mile_diff[mode] = 0
        for submode in TRANSIT_AGGREGATIONS[mode]:
            transit_mile_diff[mode] += transit_mile_diff[submode]
    ws = workbook["Tuottajahyodyt"]
    cols = CELL_INDICES["transit_miles"]["cols"]
    rows = CELL_INDICES["transit_miles"]["rows"][year]
    for mode in rows:
        for imp_type in cols:
            ws[cols[imp_type]+rows[mode]] = transit_mile_diff[imp_type][mode]
    log.info("Mileage differences calculated")

    # Calculate gains and revenues
    for tp in ["aht", "pt", "iht"]:
        data = {
            "scen_1": MatrixData(os.path.join(scenario_1, "Matrices")),
            "scen_0": MatrixData(os.path.join(scenario_0, "Matrices")),
        }
        revenues_transit = 0
        revenues_car = 0
        cols = CELL_INDICES["gains"]["cols"]
        rows = CELL_INDICES["gains"]["rows"][year]
        for tc in param.transport_classes:
            demand = {}
            for scenario in data:
                with data[scenario].open("demand", tp) as mtx:
                    demand[scenario] = mtx[tc]
            for mtx_type in ["time", "cost", "dist"]:
                cost = {scenario: read_costs(data[scenario], tp, tc, mtx_type)
                    for scenario in data}
                gains_existing, gains_additional = calc_gains(demand, cost)
                ws = workbook[TRANSLATIONS[tc]]
                ws[cols[tp]+rows[mtx_type][0]] = gains_existing
                ws[cols[tp]+rows[mtx_type][1]] = gains_additional
                if mtx_type == "cost":
                    revenue = calc_revenue(demand, cost)
                    if tc in param.transit_classes:
                        revenues_transit += revenue
                    if tc in param.assignment_modes:
                        revenues_car += revenue
        ws = workbook["Tuottajahyodyt"]
        rows = CELL_INDICES["transit_revenue"]["rows"][year]
        ws[cols[tp]+rows] = revenues_transit
        ws = workbook["Julkistaloudelliset"]
        cols = CELL_INDICES["car_revenue"]["cols"]
        rows = CELL_INDICES["car_revenue"]["rows"][year]
        ws[cols[tp]+rows] = revenues_car
        log.info("Gains and revenues calculated for {}".format(tp))
    log.info("Year {} completed".format(year))


def read_miles(scenario_path):
    """Read vehicle km data from file."""
    file_path = os.path.join(scenario_path, VEHICLE_KMS_FILE)
    return pandas.read_csv(file_path, delim_whitespace=True)


def read_transit_miles(scenario_path):
    """Read transit vehicle travel time and dist data from file."""
    file_path = os.path.join(scenario_path, TRANSIT_KMS_FILE)
    return pandas.read_csv(file_path, delim_whitespace=True)


def read_costs(matrixdata, time_period, transport_class, mtx_type):
    mtx_label = transport_class.split('_')[0]
    ass_class = mtx_label if mtx_label == "bike" else transport_class
    if mtx_label == "bike" and mtx_type == "cost":
        matrix = 0
    else:
        with matrixdata.open(mtx_type, time_period) as mtx:
            matrix = mtx[ass_class]
            zone_numbers = mtx.zone_numbers
    if transport_class == "transit_work" and mtx_type == "cost":
        nr_trips = numpy.full_like(
            matrix, TRANSIT_TRIPS_PER_MONTH["work_capital_region"])
        surrounding = numpy.searchsorted(
            zone_numbers, zone_param.areas["surrounding"][0])
        nr_trips[surrounding:, :] = TRANSIT_TRIPS_PER_MONTH["work_surrounding"]
        nr_trips = 0.5 * (nr_trips+nr_trips.T)
        matrix /= nr_trips
    elif transport_class == "transit_leisure" and mtx_type == "cost":
        matrix /= TRANSIT_TRIPS_PER_MONTH["leisure"]
    return matrix


def calc_gains(demands, costs):
    """Calculate difference in consumer surplus between scen_1 and scen_0.

    Parameters
    ----------
    demands : dict
        scen_0 : numpy.ndarray
            Demand matrix for scenario 0
        scen_1 : numpy.ndarray
            Demand matrix for scenario 1
    costs : dict
        scen_0 : numpy.ndarray
            Impedance matrix for scenario 0
        scen_1 : numpy.ndarray
            Impedance matrix for scenario 1

    Returns
    -------
    float
        Calculated gain for existing users
    float
        Calculated gain for new or evicted users
    """
    gain = costs["scen_1"] - costs["scen_0"]
    demand_change = demands["scen_1"] - demands["scen_0"]
    gains_existing = ((demands["scen_0"]*gain)[demand_change >= 0].sum()
                      + (demands["scen_1"]*gain)[demand_change < 0].sum())
    gains_additional = (0.5*(demand_change*gain)[demand_change >= 0].sum()
                        - 0.5*(demand_change*gain)[demand_change < 0].sum())
    return gains_existing, gains_additional


def calc_revenue(demands, costs):
    """Calculate difference in producer revenue between scen_1 and scen_0.

    Parameters
    ----------
    demands : dict
        scen_0 : numpy.ndarray
            Demand matrix for scenario 0
        scen_1 : numpy.ndarray
            Demand matrix for scenario 1
    costs : dict
        scen_0 : numpy.ndarray
            Impedance matrix for scenario 0
        scen_1 : numpy.ndarray
            Impedance matrix for scenario 1

    Returns
    -------
    float
        Calculated revenue
    """
    demand_change = demands["scen_1"] - demands["scen_0"]
    cost_change = costs["scen_1"] - costs["scen_0"]
    revenue = ((costs["scen_1"]*demand_change)[demand_change >= 0].sum()
               + (cost_change*demands["scen_0"])[demand_change >= 0].sum()
               + (costs["scen_0"]*demand_change)[demand_change < 0].sum()
               + (cost_change*demands["scen_1"])[demand_change < 0].sum())
    return revenue


if __name__ == "__main__":
    config = Config().read_from_file()
    config.LOG_FORMAT = "JSON"
    config.SCENARIO_NAME = "cba"
    parser = ArgumentParser(epilog="Calculates the Cost-Benefit Analysis between Results of two HELMET-Scenarios, "
                                   "and writes the outcome in CBA_kehikko.xlsx -file (in same folder).")
    parser.add_argument(
        "baseline_scenario", type=str, help="A 'do-nothing' baseline scenario")
    parser.add_argument(
        "projected_scenario", type=str,
        help="A projected scenario, compared to the baseline scenario")
    parser.add_argument(
        "baseline_scenario_2", nargs='?', type=str,
        help="A 'do-nothing' baseline scenario for second forecast year (optional)")
    parser.add_argument(
        "projected_scenario_2", nargs='?', type=str,
        help="A projected scenario, compared to the baseline scenario for second forecast year (optional)")
    parser.add_argument(
        "--results-path", dest="results_path", type=str, required=True,
        help="Path to Results directory.")
    args = parser.parse_args()
    log.initialize(config)
    wb = load_workbook(os.path.join(SCRIPT_DIR, "CBA_kehikko.xlsx"))
    run_cost_benefit_analysis(
        args.baseline_scenario, args.projected_scenario, 1, wb)
    if args.baseline_scenario_2 is not None and args.baseline_scenario_2 != "undefined":
        run_cost_benefit_analysis(
            args.baseline_scenario_2, args.projected_scenario_2, 2, wb)
    results_filename = "cba_{}_{}.xlsx".format(
        os.path.basename(args.projected_scenario),
        os.path.basename(args.baseline_scenario))
    wb.save(os.path.join(args.results_path, results_filename))
    log.info("CBA results saved to file: {}".format(results_filename))
