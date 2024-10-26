from __future__ import annotations

from typing import Any, Dict
import pandas
from pathlib import Path
try:
    from openpyxl import Workbook, load_workbook
    _use_txt = False
except ImportError:
    _use_txt = True


class ResultsData:
    """
    Saves all result data to same folder.
    """
    def __init__(self, results_directory_path: Path):
        self.path = results_directory_path
        self.path.mkdir(parents=True, exist_ok=True)
        self._line_buffer: Dict[str, Any] = {}
        self._df_buffer: Dict[str, Any] = {}
        self._xlsx_buffer: Dict[str, Any] = {}

    def flush(self):
        """Save to files and empty buffers."""
        for filename, buffer in self._line_buffer.items():
            buffer.close()
        self._line_buffer = {}

        for filename, df in self._df_buffer.items():
            df.to_csv(self.path / filename, sep='\t', float_format="%1.5f")
        self._df_buffer = {}

        for filename, wb in self._xlsx_buffer.items():
            wb.save(self.path / f"{filename}.xlsx")
        self._xlsx_buffer = {}

    def print_data(self, data: pandas.Series, filename: str, colname: str):
        """Save data to DataFrame buffer (printed to text file when flushing).

        Parameters
        ----------
        data : pandas.Series
            Data to add as a new column to DataFrame
        filename : str
            Name of file where data is pushed (can contain other data)
        colname : str
            Desired name of this column
        """
        if filename not in self._df_buffer:
            self._df_buffer[filename] = pandas.DataFrame(data, columns=[colname])
        else:
            df = self._df_buffer[filename]
            self._df_buffer[filename] = df.reindex(
                df.index.union(data.index), copy=False)
            self._df_buffer[filename][colname] = data

    def print_line(self, line: str, filename: str):
        """Write text to line in file (closed when flushing).

        Parameters
        ----------
        line : str
            Row of text
        filename : str
            Name of file where text is pushed (can contain other text)
        """
        try:
            buffer = self._line_buffer[filename]
        except KeyError:
            buffer = (self.path / f"{filename}.txt").open(mode='w', encoding='utf-8')
            self._line_buffer[filename] = buffer
        buffer.write(line + "\n")

    def print_matrix(self, data: pandas.DataFrame, filename: str, sheetname: str):
        """Save 2-d matrix data to buffer (printed to file when flushing).

        Saves matrix both in Excel format and as list in text file.

        Parameters
        ----------
        data : pandas DataFrame
            Data to add as a new sheet to WorkBook
        filename : str
            Name of file where data is pushed (without file extension)
        sheetname : str
            Desired name of excel sheet
        """
        if _use_txt:
            # If no Workbook module available (= _use_txt), save data to csv
            data.to_csv(
                self.path / f"{filename}_{sheetname}.txt",
                sep='\t', float_format="%8.1f")
        else:
            # Get/create new worksheet
            if filename in self._xlsx_buffer:
                ws = self._xlsx_buffer[filename].create_sheet(sheetname)
            else:
                self._xlsx_buffer[filename] = Workbook()
                ws = self._xlsx_buffer[filename].active
                ws.title = sheetname
            # Write data to each cell
            for j in range(0, data.shape[1]):
                ws.cell(row=1, column=j+2).value = data.columns[j]
            for i in range(0, data.shape[0]):
                ws.cell(row=i+2, column=1).value = data.index[i]
                for j in range(0, data.shape[1]):
                    ws.cell(row=i+2, column=j+2).value = data.iloc[i, j]
        # Create text file
        sheetname = sheetname.replace("_", "\t")
        for j in data.columns:
            for i in data.index:
                self.print_line(
                    "{}\t{}\t{}\t{}".format(i, j, sheetname, str(data[j][i])),
                    filename)
